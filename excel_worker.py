import datetime
import math
import zipfile
import openpyxl
import openpyxl.utils.exceptions
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side, Font
from tkinter import messagebox
import re
from efc.interfaces.iopenpyxl import OpenpyxlInterface

import db_worker

columns_names_row = 4
values_first_row = 6

ignore_positions = ["Нулевая статья"]

first_direction_name = "Связанные расчеты"
first_direction = 2
second_direction_names = ["Прямые платежи", "На расчетный счет"]
second_direction = 1
default_direction = second_direction

comment_author_default = "Unknown"

arithmetic_regexp = re.compile(r"[^\-+*/^()=0-9.\s]")
excel_date_regexp = re.compile(r"20\d\d-(0[1-9]|1[012])-(0[1-9]|1[0-9]|2[0-9]|3[01]) 00:00:00")


# Является ли объект числом
def is_number(obj):
    try:
        float(obj)
        return True
    except ValueError:
        return False


# Округлить число до сотых (в большую сторону)
def round_to_cents(value):
    return math.ceil(value * 100.0) / 100.0


# Форматирование значения из Excel для записи в БД
def format_value_to_db(value, calculator, sheet_name, row, column):
    cell = openpyxl.utils.get_column_letter(column) + str(row)
    # Проверка, является ли значение числом или формулой, результат которой - число
    try:
        if not is_number(calculator.calc_cell(cell, sheet_name)):
            return None
    except:
        return None
    # Если число - округляем его до сотых
    if is_number(value):
        return str(round_to_cents(float(value)))
    # Если не арифметическая формула - считаем ее результат, округляем до сотых
    if re.search(arithmetic_regexp, value) is not None:
        return str(round_to_cents(float(calculator.calc_cell(cell, sheet_name))))
    # Если арифметическая формула - возвращаем ее, как есть
    return value


# Форматирование значения из БД для записи в Excel
def format_value_to_excel(value):
    if value is None:
        return None
    # Если число - преобразуем к числу
    if is_number(value):
        return float(value)
    # Если не является арифметической формулой - ошибка
    if re.search(arithmetic_regexp, value) is not None:
        return None
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1).value = value
    calculator = OpenpyxlInterface(workbook)
    # Если не можем посчитать арифметическую формулу - ошибка, если можем - возвращаем, как есть
    try:
        if is_number(calculator.calc_cell("A1", sheet.title)):
            return value
        return None
    except:
        return None


# Конвертируем значение в число (применяется для свода)
def count_value(value):
    if value is None:
        return None
    # Если число - приводим к числу
    if is_number(value):
        return float(value)
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1).value = value
    calculator = OpenpyxlInterface(workbook)
    # Если не можем посчитать - ошибка, иначе возвращаем результат
    try:
        if is_number(calculator.calc_cell("A1", sheet.title)):
            return round_to_cents(float(calculator.calc_cell("A1", sheet.title)))
        return None
    except:
        return None


# Суммируем значения для записи в Excel (применяется для свода)
def sum_values_to_excel(values):
    summary = 0
    if len(values) == 0:
        return ""
    for value in values:
        if count_value(value) is None:
            return None
        summary += count_value(value)
    return summary


# Проверить корректность комментария (отметаем примечания)
def is_comment_valid(comment):
    if re.fullmatch(r"tc={[0-9A-F]{8}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{12}}", comment.author) is not None:
        return False
    return True


# Формируем итоговый комментарий (применяется для свода)
def form_summary_info(departments_info):
    summary = ""
    for department_info in departments_info:
        if department_info[-1] is None:
            department_info.pop()
        department_summary = " ".join(department_info)
        summary = f"{summary}{department_summary},\n"
    return summary[:-2]


# Для кода находим его "родительский" (используется для определения направления статьи)
def get_parent_code(code):
    code_parts = str(code).split(".")
    code_parts.reverse()
    for i in range(len(code_parts)):
        if re.search(r"[1-9]", code_parts[i]) is not None:
            code_parts[i] = "0" * len(code_parts[i])
            break
    code_parts.reverse()
    return ".".join(code_parts)


# Выполнить 1 режим - заполнение шаблона значениями из БД
def mode1_execute(infile, connection, active_codes, department_id, outfile):
    # Открываем шаблон
    try:
        workbook = openpyxl.load_workbook(infile)
    except FileNotFoundError:
        messagebox.showerror("Ошибка", f"Файл {infile} не найден")
        return False
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Ошибка", f"Файл {infile} не является файлом формата xlsx")
        return False
    except PermissionError:
        messagebox.showerror("Ошибка", f"Нет доступа к файлу {infile}")
        return False
    except (zipfile.BadZipfile, OSError):
        messagebox.showerror("Ошибка", f"Файл {infile} не удается открыть. Он поврежден или не является xlsx-файлом")
        return False
    except Exception as e:
        messagebox.showerror("Ошибка", f"Файл {infile} не удается открыть. Ошибка:\n" + str(e))
        return False

    for sheet in workbook.worksheets:
        # Определяем последний столбец даты
        max_date_column = 3
        while re.fullmatch(excel_date_regexp, str(sheet.cell(columns_names_row, max_date_column).value)) is not None:
            max_date_column += 1
        max_date_column -= 1

        # Считываем из шаблона коды статей, запоминая их строки
        active_codes_rows = {}
        all_codes_rows = {}
        currency = sheet.title
        for row in range(values_first_row, sheet.max_row + 1):
            code = sheet.cell(row, 1).value
            if code is None or ignore_positions.__contains__(sheet.cell(row, 2).value):
                continue
            if all_codes_rows.get(code) is None:
                all_codes_rows[code] = [row]
            else:
                if len(all_codes_rows[code]) == 1:
                    all_codes_rows[code] = [all_codes_rows[code][0], row]
                else:
                    messagebox.showerror("Ошибка", f"Неверный формат: статья {code} содержится в листе {currency} "
                                                   f"файла {infile} более 2 раз")
                    return False
            if active_codes.__contains__(code):
                if active_codes_rows.get(code) is None:
                    active_codes_rows[code] = [row]
                else:
                    active_codes_rows[code] = [active_codes_rows[code][0], row]
        if len(active_codes_rows) == 0:
            messagebox.showinfo("Выполнение остановлено",
                                f"Лист {currency} файла {infile} не содержит строк активных статей. "
                                f"Возможно, он не соответствует формату платежного календаря")
            return False

        # Для каждой статьи определяем направления (или направление, если строка встречается 1 раз)
        for code in active_codes_rows.keys():
            directions_rows = []
            if len(active_codes_rows.get(code)) == 2:
                directions_rows.append([first_direction, active_codes_rows.get(code)[0]])
                directions_rows.append([second_direction, active_codes_rows.get(code)[1]])
            else:
                row = active_codes_rows.get(code)[0]
                if sheet.cell(row + 1, 1).value is None and sheet.cell(row + 2, 1).value is None and \
                        sheet.cell(row + 1, 2).value == first_direction_name and \
                        second_direction_names.__contains__(sheet.cell(row + 2, 2).value):
                    directions_rows.append([first_direction, row + 1])
                    directions_rows.append([second_direction, row + 2])
                else:
                    parent_row = all_codes_rows.get(get_parent_code(code))[0]
                    direction = default_direction
                    if parent_row is not None:
                        empty_codes_rows = []
                        for r in range(parent_row, row):
                            if sheet.cell(r, 1).value is None:
                                empty_codes_rows.append(sheet.cell(r, 2).value)
                        if len(empty_codes_rows) == 1 and empty_codes_rows[0] == first_direction_name:
                            direction = first_direction
                        else:
                            if len(empty_codes_rows) == 2 and empty_codes_rows[0] == first_direction_name and \
                                    second_direction_names.__contains__(empty_codes_rows[1]):
                                direction = second_direction
                            else:
                                direction = default_direction
                    directions_rows.append([direction, row])

            # Записываем в Excel полученные из БД значения
            for dr in directions_rows:
                for column in range(3, max_date_column + 1):
                    record = db_worker.get_record(connection, department_id, code,
                                                  sheet.cell(columns_names_row, column).value.date(), dr[0], currency)
                    if record is None:
                        return False
                    if record.value is None:
                        continue
                    value = format_value_to_excel(record.value)
                    if value is not None:
                        sheet.cell(dr[1], column).value = value
                    else:
                        messagebox.showerror("Ошибка", f"В таблице {db_worker.records_table} "
                                                       f"в базе данных неверные записи")
                        return False
                    if record.comment_text is not None:
                        sheet.cell(dr[1], column).comment = Comment(record.comment_text, comment_author_default)

    # Сохраняем полученную таблицу
    try:
        workbook.save(outfile)
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Ошибка", f"Файл {outfile} не является файлом формата xlsx")
        return False
    except PermissionError:
        messagebox.showerror("Ошибка", f"Нет доступа к файлу {outfile}. Если файл открыт - закройте его")
        return False
    except Exception as e:
        messagebox.showerror("Ошибка", f"Файл {outfile} не удается открыть на запись. Ошибка:\n" + str(e))
        return False
    return True


# Выполнить 2 режим - сохранение ПК в БД
def mode2_execute(infile, active_codes):
    # Открываем ПК
    try:
        workbook = openpyxl.load_workbook(infile)
    except FileNotFoundError:
        messagebox.showerror("Ошибка", f"Файл {infile} не найден")
        return None
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Ошибка", f"Файл {infile} не является файлом формата xlsx")
        return None
    except PermissionError:
        messagebox.showerror("Ошибка", f"Нет доступа к файлу {infile}")
        return None
    except (zipfile.BadZipfile, OSError):
        messagebox.showerror("Ошибка", f"Файл {infile} не удается открыть. Он поврежден или не является xlsx-файлом")
        return None
    except Exception as e:
        messagebox.showerror("Ошибка", f"Файл {infile} не удается открыть. Ошибка:\n" + str(e))
        return None

    calc = OpenpyxlInterface(workbook)
    records = []
    for sheet in workbook.worksheets:
        # Определяем последний столбец даты
        max_date_column = 3
        while re.fullmatch(excel_date_regexp, str(sheet.cell(columns_names_row, max_date_column).value)) is not None:
            max_date_column += 1
        max_date_column -= 1

        # Считываем из ПК коды статей, запоминая их строки
        active_codes_rows = {}
        all_codes_rows = {}
        currency = sheet.title
        for row in range(values_first_row, sheet.max_row + 1):
            code = sheet.cell(row, 1).value
            if code is None or ignore_positions.__contains__(sheet.cell(row, 2).value):
                continue
            if all_codes_rows.get(code) is None:
                all_codes_rows[code] = [row]
            else:
                if len(all_codes_rows[code]) == 1:
                    all_codes_rows[code] = [all_codes_rows[code][0], row]
                else:
                    messagebox.showerror("Ошибка", f"Неверный формат: статья {code} содержится в листе {currency} "
                                                   f"файла {infile} более 2 раз")
                    return None
            if active_codes.__contains__(code):
                if active_codes_rows.get(code) is None:
                    active_codes_rows[code] = [row]
                else:
                    active_codes_rows[code] = [active_codes_rows[code][0], row]
        if len(active_codes_rows) == 0:
            messagebox.showinfo("Выполнение остановлено",
                                f"Лист {currency} файла {infile} не содержит строк активных статей. "
                                f"Возможно, он не соответствует формату платежного календаря")
            return None

        # Для каждой статьи определяем направления (или направление, если строка встречается 1 раз)
        for code in active_codes_rows.keys():
            directions_rows = []
            if len(active_codes_rows.get(code)) == 2:
                directions_rows.append([first_direction, active_codes_rows.get(code)[0]])
                directions_rows.append([second_direction, active_codes_rows.get(code)[1]])
            else:
                row = active_codes_rows.get(code)[0]
                if sheet.cell(row + 1, 1).value is None and sheet.cell(row + 2, 1).value is None and \
                        sheet.cell(row + 1, 2).value == first_direction_name and \
                        second_direction_names.__contains__(sheet.cell(row + 2, 2).value):
                    directions_rows.append([first_direction, row + 1])
                    directions_rows.append([second_direction, row + 2])
                else:
                    parent_row = all_codes_rows.get(get_parent_code(code))[0]
                    direction = default_direction
                    if parent_row is not None:
                        empty_codes_rows = []
                        for r in range(parent_row, row):
                            if sheet.cell(r, 1).value is None:
                                empty_codes_rows.append(sheet.cell(r, 2).value)
                        if len(empty_codes_rows) == 1 and empty_codes_rows[0] == first_direction_name:
                            direction = first_direction
                        else:
                            if len(empty_codes_rows) == 2 and empty_codes_rows[0] == first_direction_name and \
                                    second_direction_names.__contains__(empty_codes_rows[1]):
                                direction = second_direction
                            else:
                                direction = default_direction
                    directions_rows.append([direction, row])

            # Сохраняем значения из каждой ячейки, создавая для каждого значения запись DbWriteRecord
            for dr in directions_rows:
                for column in range(3, max_date_column + 1):
                    if sheet.cell(dr[1], column).value is not None:
                        value = format_value_to_db(str(sheet.cell(dr[1], column).value), calc, currency, dr[1], column)
                        if value is not None:
                            comment = sheet.cell(dr[1], column).comment
                            if comment is None or not is_comment_valid(comment):
                                records.append(
                                    db_worker.DbWriteRecord(code,
                                                            str(sheet.cell(columns_names_row, column).value.date()),
                                                            dr[0], currency, value))
                            else:
                                records.append(
                                    db_worker.DbWriteRecord(code,
                                                            str(sheet.cell(columns_names_row, column).value.date()),
                                                            dr[0], currency, value, comment.text))
                        else:
                            messagebox.showerror("Ошибка", f"Неверное значение: лист {currency}, строка {dr[1]}, "
                                                           f"столбец {openpyxl.utils.get_column_letter(column)}")
                            return None
    return records


# Выполнить 3 режим - свод данных от подразделений
def mode3_execute(infile, connection, active_codes, departments, outfile):
    # Открываем шаблон
    try:
        workbook = openpyxl.load_workbook(infile)
    except FileNotFoundError:
        messagebox.showerror("Ошибка", f"Файл {infile} не найден")
        return False
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Ошибка", f"Файл {infile} не является файлом формата xlsx")
        return False
    except PermissionError:
        messagebox.showerror("Ошибка", f"Нет доступа к файлу {infile}")
        return False
    except (zipfile.BadZipfile, OSError):
        messagebox.showerror("Ошибка", f"Файл {infile} не удается открыть. Он поврежден или не является xlsx-файлом")
        return False
    except Exception as e:
        messagebox.showerror("Ошибка", f"Файл {infile} не удается открыть. Ошибка:\n" + str(e))
        return False

    for sheet in workbook.worksheets:
        # Определяем последний столбец даты
        max_date_column = 3
        while re.fullmatch(excel_date_regexp, str(sheet.cell(columns_names_row, max_date_column).value)) is not None:
            max_date_column += 1
        max_date_column -= 1

        # Считываем из шаблона коды статей, запоминая их строки
        active_codes_rows = {}
        all_codes_rows = {}
        currency = sheet.title
        for row in range(values_first_row, sheet.max_row + 1):
            code = sheet.cell(row, 1).value
            if code is None or ignore_positions.__contains__(sheet.cell(row, 2).value):
                continue
            if all_codes_rows.get(code) is None:
                all_codes_rows[code] = [row]
            else:
                if len(all_codes_rows[code]) == 1:
                    all_codes_rows[code] = [all_codes_rows[code][0], row]
                else:
                    messagebox.showerror("Ошибка",
                                         f"Неверный формат: статья {code} содержится в листе {currency} "
                                         f"файла {infile} более 2 раз")
                    return False
            if active_codes.__contains__(code):
                if active_codes_rows.get(code) is None:
                    active_codes_rows[code] = [row]
                else:
                    active_codes_rows[code] = [active_codes_rows[code][0], row]
        if len(active_codes_rows) == 0:
            messagebox.showinfo("Выполнение остановлено",
                                f"Лист {currency} файла {infile} не содержит строк активных статей. "
                                f"Возможно, он не соответствует формату платежного календаря")
            return False

        # Для каждой статьи определяем направления (или направление, если строка встречается 1 раз)
        for code in active_codes_rows.keys():
            directions_rows = []
            if len(active_codes_rows.get(code)) == 2:
                directions_rows.append([first_direction, active_codes_rows.get(code)[0]])
                directions_rows.append([second_direction, active_codes_rows.get(code)[1]])
            else:
                row = active_codes_rows.get(code)[0]
                if sheet.cell(row + 1, 1).value is None and sheet.cell(row + 2, 1).value is None and \
                        sheet.cell(row + 1, 2).value == first_direction_name and \
                        second_direction_names.__contains__(sheet.cell(row + 2, 2).value):
                    directions_rows.append([first_direction, row + 1])
                    directions_rows.append([second_direction, row + 2])
                else:
                    parent_row = all_codes_rows.get(get_parent_code(code))[0]
                    direction = default_direction
                    if parent_row is not None:
                        empty_codes_rows = []
                        for r in range(parent_row, row):
                            if sheet.cell(r, 1).value is None:
                                empty_codes_rows.append(sheet.cell(r, 2).value)
                        if len(empty_codes_rows) == 1 and empty_codes_rows[0] == first_direction_name:
                            direction = first_direction
                        else:
                            if len(empty_codes_rows) == 2 and empty_codes_rows[0] == first_direction_name and \
                                    second_direction_names.__contains__(empty_codes_rows[1]):
                                direction = second_direction
                            else:
                                direction = default_direction
                    directions_rows.append([direction, row])
            # Получаем для каждой ячейки значения от всех подразделений, суммируем и записываем в таблицу
            for dr in directions_rows:
                for column in range(3, max_date_column + 1):
                    values = []
                    departments_info = []
                    for department in departments:
                        record = db_worker.get_record(connection, department, code,
                                                      sheet.cell(columns_names_row, column).value.date(), dr[0],
                                                      currency)
                        if record is None:
                            return False
                        if record.value is None:
                            continue
                        values.append(record.value)
                        department_name = db_worker.get_department_name(connection, department)
                        if department_name is None:
                            return False
                        departments_info.append([department_name, record.value, record.comment_text])
                    if len(values) == 0:
                        continue
                    value = sum_values_to_excel(values)
                    summary_comment = form_summary_info(departments_info)
                    if value is not None:
                        sheet.cell(dr[1], column).value = value
                    else:
                        messagebox.showerror("Ошибка", f"В таблице {db_worker.records_table} "
                                                       f"в базе данных неверные записи")
                        return False
                    if summary_comment != "":
                        sheet.cell(dr[1], column).comment = Comment(summary_comment, comment_author_default, 120, 250)

    # Сохраняем полученную таблицу
    try:
        workbook.save(outfile)
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Ошибка", f"Файл {outfile} не является файлом формата xlsx")
        return False
    except PermissionError:
        messagebox.showerror("Ошибка", f"Нет доступа к файлу {outfile}. Если файл открыт - закройте его")
        return False
    except Exception as e:
        messagebox.showerror("Ошибка", f"Файл {outfile} не удается открыть на запись. Ошибка:\n" + str(e))
        return False
    return True


# Выполнить 4 режим - выгрузка всех записей из БД в простую таблицу
def mode4_execute(records, outfile):
    # Формируем таблицу
    workbook = Workbook()
    sheet = workbook.active
    row = 1
    sheet.cell(row, 1).value = "Выгрузка от:"
    sheet.cell(row, 2).value = datetime.datetime.now()
    row += 1
    sheet.cell(row, 1).value = "id"
    sheet.cell(row, 2).value = "Подразделение"
    sheet.cell(row, 3).value = "Статья"
    sheet.cell(row, 4).value = "Валюта"
    sheet.cell(row, 5).value = "Направление"
    sheet.cell(row, 6).value = "Дата"
    sheet.cell(row, 7).value = "Сумма"
    sheet.cell(row, 8).value = "Заметка"
    sheet.cell(row, 9).value = "Отметка времени"
    sheet.column_dimensions['A'].width = 11
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 17
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    for column in range(9):
        sheet.cell(row, column + 1).border = Border(top=Side(border_style='medium'), right=Side(border_style='medium'),
                                                    bottom=Side(border_style='medium'),
                                                    left=Side(border_style='medium'))
        sheet.cell(row, column + 1).font = Font(bold=True)
    for record in records:
        if len(record) != 9:
            messagebox.showerror("Ошибка", "Записи из базы данных прочитаны неверно неверно")
            return False
        row += 1
        sheet.cell(row, 1).value = record[0]
        sheet.cell(row, 2).value = record[1]
        sheet.cell(row, 3).value = record[2]
        sheet.cell(row, 4).value = record[3]
        if record[4] == first_direction:
            sheet.cell(row, 5).value = first_direction_name
        else:
            if record[4] == second_direction:
                if record[2][0] == '1':
                    sheet.cell(row, 5).value = second_direction_names[0]
                else:
                    sheet.cell(row, 5).value = second_direction_names[1]
            else:
                messagebox.showerror("Ошибка", f"Неверное направление {record[4]} у записи {record[0]}")
                return False
        if isinstance(record[5], datetime.datetime):
            sheet.cell(row, 6).value = record[5].date()
        else:
            messagebox.showerror("Ошибка", f"Дата записи {record[0]} не является датой со временем")
            return False
        if is_number(record[6]):
            sheet.cell(row, 7).value = float(record[6])
        else:
            sheet.cell(row, 7).value = record[6]
        sheet.cell(row, 8).value = record[7]
        if isinstance(record[8], datetime.datetime):
            sheet.cell(row, 9).value = datetime.datetime(record[8].year, record[8].month, record[8].day,
                                                         record[8].hour, record[8].minute, record[8].second)
        else:
            messagebox.showerror("Ошибка", f"Отметка времени записи {record[0]} не является датой со временем")
            return False
    sheet.title = "Выгрузка БД"

    # Сохраняем полученную таблицу
    try:
        workbook.save(outfile)
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Ошибка", f"Файл {outfile} не является файлом формата xlsx")
        return False
    except PermissionError:
        messagebox.showerror("Ошибка", f"Нет доступа к файлу {outfile}. Если файл открыт - закройте его")
        return False
    except Exception as e:
        messagebox.showerror("Ошибка", f"Файл {outfile} не удается открыть на запись. Ошибка:\n" + str(e))
        return False
    return True
